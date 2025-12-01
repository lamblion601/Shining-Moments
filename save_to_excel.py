import pandas as pd
from datetime import datetime

# 크롤링한 데이터라고 가정한 샘플 데이터
# 실제로는 크롤링한 데이터를 여기에 넣으면 됩니다
def create_sample_data():
    """
    크롤링한 데이터라고 가정한 샘플 데이터를 생성합니다.
    실제 크롤링 코드에서는 이 부분을 크롤링 결과로 대체하세요.
    """
    data = {
        '제목': [
            '제품 A 리뷰',
            '제품 B 리뷰',
            '제품 C 리뷰',
            '제품 D 리뷰',
            '제품 E 리뷰'
        ],
        '평점': [4.5, 3.8, 5.0, 4.2, 4.7],
        '가격': [29900, 15900, 45900, 32900, 21900],
        '리뷰수': [1250, 890, 2340, 1560, 670],
        '작성일': [
            '2025-01-15',
            '2025-01-14',
            '2025-01-13',
            '2025-01-12',
            '2025-01-11'
        ]
    }
    
    # DataFrame 생성
    df = pd.DataFrame(data)
    
    return df


def save_to_excel(df, filename='크롤링데이터.xlsx', sheet_name='데이터'):
    """
    pandas DataFrame을 엑셀 파일로 저장합니다.
    
    Parameters:
    df: pandas DataFrame - 저장할 데이터
    filename: str - 저장할 엑셀 파일명 (기본값: '크롤링데이터.xlsx')
    sheet_name: str - 시트 이름 (기본값: '데이터')
    """
    try:
        # 엑셀 파일로 저장
        # index=False는 인덱스를 엑셀에 포함하지 않음
        # engine='openpyxl'은 xlsx 형식을 위한 엔진
        df.to_excel(filename, sheet_name=sheet_name, index=False, engine='openpyxl')
        
        print(f"[성공] 데이터가 '{filename}' 파일로 성공적으로 저장되었습니다!")
        print(f"  - 총 {len(df)}개의 행이 저장되었습니다.")
        print(f"  - 컬럼: {', '.join(df.columns.tolist())}")
        
    except Exception as e:
        print(f"[오류] 저장 중 오류가 발생했습니다: {e}")


def save_to_excel_advanced(df, filename='크롤링데이터_상세.xlsx'):
    """
    여러 시트와 포맷팅이 포함된 고급 엑셀 저장 예제
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    try:
        # 기본 저장
        df.to_excel(filename, sheet_name='전체데이터', index=False, engine='openpyxl')
        
        # 엑셀 파일 열기 (포맷팅을 위해)
        wb = load_workbook(filename)
        ws = wb['전체데이터']
        
        # 헤더 스타일 설정
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        # 헤더 행에 스타일 적용
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 컬럼 너비 자동 조정 (대략적인 예시)
        column_widths = {
            'A': 25,  # 제목
            'B': 12,  # 평점
            'C': 15,  # 가격
            'D': 12,  # 리뷰수
            'E': 15   # 작성일
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 요약 시트 추가
        summary_data = {
            '항목': ['총 데이터 수', '평균 평점', '평균 가격', '총 리뷰수'],
            '값': [
                len(df),
                round(df['평점'].mean(), 2),
                round(df['가격'].mean(), 0),
                df['리뷰수'].sum()
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        
        # 요약 시트 추가
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
            summary_df.to_excel(writer, sheet_name='요약', index=False)
        
        print(f"[성공] 상세 포맷팅이 적용된 '{filename}' 파일이 생성되었습니다!")
        print(f"  - 시트: '전체데이터', '요약'")
        
    except Exception as e:
        print(f"[오류] 저장 중 오류가 발생했습니다: {e}")


if __name__ == "__main__":
    print("=" * 50)
    print("크롤링 데이터를 엑셀로 저장하는 예제")
    print("=" * 50)
    
    # 샘플 데이터 생성 (실제로는 크롤링한 데이터를 사용)
    print("\n[1단계] 데이터 준비 중...")
    df = create_sample_data()
    print(f"[완료] {len(df)}개의 데이터 행을 준비했습니다.")
    print("\n데이터 미리보기:")
    print(df.head())
    
    # 기본 엑셀 저장
    print("\n[2단계] 엑셀 파일로 저장 중...")
    save_to_excel(df, filename='크롤링데이터.xlsx')
    
    # 고급 엑셀 저장 (포맷팅 포함)
    print("\n[3단계] 포맷팅이 포함된 엑셀 파일 생성 중...")
    save_to_excel_advanced(df, filename='크롤링데이터_상세.xlsx')
    
    print("\n" + "=" * 50)
    print("완료! 두 개의 엑셀 파일이 생성되었습니다.")
    print("=" * 50)

